"""Write renderer plans to .xlsx workbooks."""

from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import range_boundaries

from .renderer import CellFormat, RenderPlan


def write_xlsx(plan: RenderPlan, output_path: str) -> None:
    """Persist a RenderPlan as a formatted .xlsx workbook."""

    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = "Sheet"

    for setup in plan.sheet_setups:
        worksheet = workbook.create_sheet(title=setup.sheet)
        for column, width in setup.column_widths.items():
            worksheet.column_dimensions[column].width = width
        if setup.freeze_panes:
            worksheet.freeze_panes = setup.freeze_panes

    for write in plan.writes:
        worksheet = _worksheet_for(workbook, write.sheet)
        worksheet[write.cell].value = write.value

    for cell_format in plan.formats:
        worksheet = _worksheet_for(workbook, cell_format.sheet)
        _apply_format(worksheet, cell_format)

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        workbook.remove(workbook["Sheet"])

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _worksheet_for(workbook: Workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    return workbook.create_sheet(title=sheet_name)


def _apply_format(worksheet, cell_format: CellFormat) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(cell_format.range)
    font_color = _normalize_font_color(cell_format.font_color)

    for row in worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            if cell_format.number_format is not None:
                cell.number_format = cell_format.number_format
            fill_color = _normalize_color(cell_format.fill_color, field="fill_color")
            if fill_color is not None:
                cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)

            top_border_color = _normalize_color(
                cell_format.top_border_color,
                field="top_border_color",
            )
            bottom_border_color = _normalize_color(
                cell_format.bottom_border_color,
                field="bottom_border_color",
            )
            if top_border_color is not None or bottom_border_color is not None:
                border = copy(cell.border) if cell.border is not None else Border()
                if top_border_color is not None:
                    border.top = Side(style="thin", color=top_border_color)
                if bottom_border_color is not None:
                    border.bottom = Side(style="thin", color=bottom_border_color)
                cell.border = border

            if cell_format.bold is None and font_color is None and cell_format.underline is None:
                continue

            font = copy(cell.font) if cell.font is not None else Font()
            if cell_format.bold is not None:
                font.bold = cell_format.bold
            if font_color is not None:
                font.color = font_color
            if cell_format.underline is not None:
                font.underline = "single" if cell_format.underline else None
            cell.font = font


def _normalize_font_color(color: str | None) -> str | None:
    return _normalize_color(color, field="font_color")


def _normalize_color(color: str | None, *, field: str) -> str | None:
    if not color:
        return None
    normalized = color.lstrip("#").upper()
    if len(normalized) == 6:
        return normalized
    if len(normalized) == 8:
        return normalized[-6:]
    raise ValueError(f"Unsupported {field}: {color}")


__all__ = ["write_xlsx"]
