from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .workbook_communication_specs import (
    _CellSpec,
    _DirectCellSpec,
    _SCENARIO_CASES,
    _SCENARIO_DIRECT_CELL_SPECS,
    _VALUATION_DIRECT_CELL_SPECS,
    _VALUATION_SUMMARY_SPECS,
    _VISUAL_CLEANUP_SPECS,
    _VisualCleanupSpec as _VisualCleanupSpec,
)
from .workbook_communication_sources import (
    WorkbookCommunicationSourceValue,
    _derive_expected_return as _derive_expected_return,
    _field_from_mapping as _field_from_mapping,
    _iter_scenarios as _iter_scenarios,
    _load_artifact_payload as _load_artifact_payload,
    _merge_current_model_values as _merge_current_model_values,
    _merge_expected_value_values as _merge_expected_value_values,
    _merge_price_target_values as _merge_price_target_values,
    _merge_scenario_values as _merge_scenario_values,
    _merge_valuation_method_values as _merge_valuation_method_values,
    _normalize_source_values,
    _scenario_case as _scenario_case,
    _set_number as _set_number,
    source_values_from_artifacts as source_values_from_artifacts,
)
from .workbook_communication_types import (
    WorkbookCommunicationCandidateWrite,
    WorkbookCommunicationFieldReport,
    WorkbookCommunicationFieldStatus,
    WorkbookCommunicationMaterializationResult,
    WorkbookCommunicationMaterializationStatus,
    WorkbookCommunicationReadiness,
    WorkbookCommunicationStatus,
)
from .workbook_communication_utils import (
    _artifact_ref as _artifact_ref,
    _first_mapping as _first_mapping,
    _first_number as _first_number,
    _get_path as _get_path,
    _is_blankish,
    _is_formula,
    _normalize_label,
    _number_from_source as _number_from_source,
    _number_or_none as _number_or_none,
    _string_or_none as _string_or_none,
    _values_equivalent,
)


def plan_workbook_communication_readiness(
    workbook_path: str | Path,
    source_values: Mapping[str, Any] | None = None,
) -> WorkbookCommunicationReadiness:
    """Plan code-owned workbook communication writes without mutating the workbook.

    The planner treats formulas without cached/static-display values as not ready
    for communication surfaces. If a typed artifact supplies the same field, the
    report emits a candidate write that a later finalization layer can apply.
    """

    path = Path(workbook_path)
    sources = _normalize_source_values(source_values or {})
    if not path.exists():
        return WorkbookCommunicationReadiness(
            workbook_path=str(path),
            workbook_exists=False,
            status="blocked",
            issues=["workbook_not_found"],
            summary={"blocked": 1},
        )

    formulas = load_workbook(path, data_only=False)
    cached = load_workbook(path, data_only=True)

    fields: list[WorkbookCommunicationFieldReport] = []
    if "Summary" not in formulas.sheetnames:
        fields.append(
            WorkbookCommunicationFieldReport(
                field="summary",
                label="Summary",
                sheet="Summary",
                status="missing_sheet",
                issue="Summary sheet is missing.",
            )
        )
    else:
        fields.extend(_summary_valuation_fields(formulas["Summary"], cached["Summary"], sources))
        fields.extend(_summary_scenario_target_fields(formulas["Summary"], cached["Summary"], sources))
    fields.extend(_direct_cell_fields(formulas, cached, sources, _VALUATION_DIRECT_CELL_SPECS))
    fields.extend(_direct_cell_fields(formulas, cached, sources, _SCENARIO_DIRECT_CELL_SPECS))

    candidate_writes = [field.candidate_write for field in fields if field.candidate_write is not None]
    required_blocked = _has_required_blockers(fields)
    cleanup_writes = [] if required_blocked else _visual_cleanup_writes(
        formulas,
        cached,
        protected_cells={(field.sheet, field.cell) for field in fields if field.cell},
    )
    candidate_writes.extend(cleanup_writes)
    summary = _summarize_fields(fields)
    if cleanup_writes:
        summary["visual_cleanup"] = len(cleanup_writes)
    issues = _issues_from_fields(fields)
    status = _readiness_status(fields, candidate_writes)
    return WorkbookCommunicationReadiness(
        workbook_path=str(path),
        status=status,
        fields=fields,
        candidate_writes=candidate_writes,
        issues=issues,
        summary=summary,
    )


def materialize_workbook_communication(
    workbook_path: str | Path,
    source_values: Mapping[str, Any] | None = None,
    *,
    output_path: str | Path | None = None,
    allow_in_place: bool = False,
) -> WorkbookCommunicationMaterializationResult:
    """Write a communication-ready workbook copy from trusted candidate writes.

    By default this writes ``<stem>.communication.xlsx`` next to the source
    workbook. It replaces only contracted communication cells that the planner
    marked as source-backed ``materializable`` and records provenance in a
    hidden ``_FMS_Communication`` ledger sheet.
    """

    path = Path(workbook_path)
    pre_readiness = plan_workbook_communication_readiness(path, source_values)
    if pre_readiness.status == "blocked":
        return WorkbookCommunicationMaterializationResult(
            workbook_path=str(path),
            status="blocked",
            pre_readiness=pre_readiness,
            issues=pre_readiness.issues or ["workbook_communication_blocked"],
        )
    if not pre_readiness.candidate_writes:
        return WorkbookCommunicationMaterializationResult(
            workbook_path=str(path),
            output_path=str(path),
            status="noop",
            pre_readiness=pre_readiness,
            post_readiness=pre_readiness,
        )

    output = Path(output_path) if output_path is not None else _default_materialized_path(path)
    if output.resolve() == path.resolve() and not allow_in_place:
        return WorkbookCommunicationMaterializationResult(
            workbook_path=str(path),
            output_path=str(output),
            status="blocked",
            pre_readiness=pre_readiness,
            issues=["output_path_matches_input; set allow_in_place=True or choose a copy path"],
        )

    workbook = load_workbook(path, data_only=False)
    ledger = _reset_communication_ledger(workbook)
    writes_applied: list[WorkbookCommunicationCandidateWrite] = []
    for write in pre_readiness.candidate_writes:
        if write.sheet not in workbook.sheetnames:
            continue
        worksheet = workbook[write.sheet]
        cell = worksheet[write.cell]
        previous_formula = cell.value if _is_formula(cell.value) else None
        cell.value = write.value
        writes_applied.append(write)
        ledger.append(
            [
                write.field,
                write.label,
                write.sheet,
                write.cell,
                write.value,
                write.source_kind,
                write.source_ref,
                previous_formula,
                write.current_display_value,
                write.reason,
            ]
        )

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    post_readiness = plan_workbook_communication_readiness(output, source_values)
    status: WorkbookCommunicationMaterializationStatus = "materialized" if writes_applied else "noop"
    return WorkbookCommunicationMaterializationResult(
        workbook_path=str(path),
        output_path=str(output),
        status=status,
        pre_readiness=pre_readiness,
        post_readiness=post_readiness,
        writes_applied=writes_applied,
        issues=[],
    )


def _summary_valuation_fields(
    formulas: Worksheet,
    cached: Worksheet,
    sources: Mapping[str, WorkbookCommunicationSourceValue],
) -> list[WorkbookCommunicationFieldReport]:
    fields: list[WorkbookCommunicationFieldReport] = []
    for spec in _VALUATION_SUMMARY_SPECS:
        row = _find_label_row(formulas, spec.label)
        if row is None:
            fields.append(
                WorkbookCommunicationFieldReport(
                    field=spec.field,
                    label=spec.label,
                    sheet=formulas.title,
                    required=spec.required,
                    status="missing_cell",
                    issue=f"Could not find Summary row labelled {spec.label!r}.",
                )
            )
            continue
        fields.append(
            _field_report(
                spec=spec,
                formulas=formulas,
                cached=cached,
                cell=f"B{row}",
                source=sources.get(spec.source_field),
            )
        )
    return fields


def _summary_scenario_target_fields(
    formulas: Worksheet,
    cached: Worksheet,
    sources: Mapping[str, WorkbookCommunicationSourceValue],
) -> list[WorkbookCommunicationFieldReport]:
    target_column = _find_table_value_column(formulas, "Case", "Target Price")
    if target_column is None:
        return []

    fields: list[WorkbookCommunicationFieldReport] = []
    for case in _SCENARIO_CASES:
        row = _find_label_row(formulas, case, max_col=1)
        if row is None:
            continue
        label = f"{case.title()} Target Price"
        spec = _CellSpec(
            f"summary.scenario.{case}.target_price",
            label,
            f"summary.scenario.{case}.target_price",
            False,
        )
        fields.append(
            _field_report(
                spec=spec,
                formulas=formulas,
                cached=cached,
                cell=f"{target_column}{row}",
                source=sources.get(spec.source_field),
            )
        )
    return fields


def _direct_cell_fields(
    formulas: Any,
    cached: Any,
    sources: Mapping[str, WorkbookCommunicationSourceValue],
    specs: Iterable[_DirectCellSpec],
) -> list[WorkbookCommunicationFieldReport]:
    fields: list[WorkbookCommunicationFieldReport] = []
    for spec in specs:
        if spec.sheet not in formulas.sheetnames:
            fields.append(
                WorkbookCommunicationFieldReport(
                    field=spec.field,
                    label=spec.label,
                    sheet=spec.sheet,
                    cell=spec.cell,
                    required=spec.required,
                    status="missing_sheet",
                    issue=f"{spec.sheet} sheet is missing.",
                )
            )
            continue
        cell_spec = _CellSpec(spec.field, spec.label, spec.source_field, spec.required)
        fields.append(
            _field_report(
                spec=cell_spec,
                formulas=formulas[spec.sheet],
                cached=cached[spec.sheet],
                cell=spec.cell,
                source=sources.get(spec.source_field),
            )
        )
    return fields


def _visual_cleanup_writes(
    formulas: Any,
    cached: Any,
    *,
    protected_cells: set[tuple[str, str]],
) -> list[WorkbookCommunicationCandidateWrite]:
    writes: list[WorkbookCommunicationCandidateWrite] = []
    for spec in _VISUAL_CLEANUP_SPECS:
        if (spec.sheet, spec.cell) in protected_cells:
            continue
        if spec.sheet not in formulas.sheetnames or spec.sheet not in cached.sheetnames:
            continue
        formula_value = formulas[spec.sheet][spec.cell].value
        cached_value = cached[spec.sheet][spec.cell].value
        if not _is_formula(formula_value) or not _is_blankish(cached_value):
            continue
        writes.append(
            WorkbookCommunicationCandidateWrite(
                field=spec.field,
                label=spec.label,
                sheet=spec.sheet,
                cell=spec.cell,
                value=None,
                source_kind="presentation_cleanup",
                current_formula=formula_value,
                current_display_value=cached_value,
                reason="communication_copy_clears_noncontracted_formula_detail",
            )
        )
    return writes


def _field_report(
    *,
    spec: _CellSpec,
    formulas: Worksheet,
    cached: Worksheet,
    cell: str,
    source: WorkbookCommunicationSourceValue | None,
) -> WorkbookCommunicationFieldReport:
    formula_value = formulas[cell].value
    cached_value = cached[cell].value
    current_formula = formula_value if _is_formula(formula_value) else None
    current_display_value = cached_value if current_formula is not None else formula_value
    has_source = source is not None and not _is_blankish(source.value)
    has_display = not _is_blankish(current_display_value)

    if has_source and (not has_display or not _values_equivalent(source.value, current_display_value)):
        reason = (
            "typed_source_can_materialize_formula_only_display"
            if not has_display
            else "typed_source_updates_stale_display"
        )
        candidate = WorkbookCommunicationCandidateWrite(
            field=spec.field,
            label=spec.label,
            sheet=formulas.title,
            cell=cell,
            value=source.value,
            source_kind=source.source_kind,
            source_ref=source.source_ref,
            current_formula=current_formula,
            current_display_value=current_display_value,
            reason=reason,
        )
        return WorkbookCommunicationFieldReport(
            field=spec.field,
            label=spec.label,
            sheet=formulas.title,
            cell=cell,
            required=spec.required,
            status="materializable",
            current_formula=current_formula,
            current_display_value=current_display_value,
            source_value=source.value,
            source_kind=source.source_kind,
            source_ref=source.source_ref,
            candidate_write=candidate,
        )

    if has_display:
        return WorkbookCommunicationFieldReport(
            field=spec.field,
            label=spec.label,
            sheet=formulas.title,
            cell=cell,
            required=spec.required,
            status="populated",
            current_formula=current_formula,
            current_display_value=current_display_value,
            source_value=source.value if source else None,
            source_kind=source.source_kind if source else None,
            source_ref=source.source_ref if source else None,
        )

    status: WorkbookCommunicationFieldStatus = "formula_only" if current_formula is not None else "missing_source"
    issue = "Formula has no cached display value." if current_formula is not None else "No display value or typed source."
    return WorkbookCommunicationFieldReport(
        field=spec.field,
        label=spec.label,
        sheet=formulas.title,
        cell=cell,
        required=spec.required,
        status=status,
        current_formula=current_formula,
        current_display_value=current_display_value,
        issue=issue,
    )


def _default_materialized_path(path: Path) -> Path:
    return path.with_name(f"{path.stem}.communication{path.suffix}")


def _reset_communication_ledger(workbook: Any) -> Any:
    title = "_FMS_Communication"
    if title in workbook.sheetnames:
        del workbook[title]
    ledger = workbook.create_sheet(title)
    ledger.sheet_state = "hidden"
    ledger.append(
        [
            "field",
            "label",
            "sheet",
            "cell",
            "value",
            "source_kind",
            "source_ref",
            "previous_formula",
            "previous_display_value",
            "reason",
        ]
    )
    return ledger


def _readiness_status(
    fields: list[WorkbookCommunicationFieldReport],
    candidate_writes: list[WorkbookCommunicationCandidateWrite],
) -> WorkbookCommunicationStatus:
    if _has_required_blockers(fields):
        return "blocked"
    if candidate_writes:
        return "actionable"
    return "ready"


def _has_required_blockers(fields: list[WorkbookCommunicationFieldReport]) -> bool:
    return any(
        field.required
        and field.status in {"formula_only", "missing_source", "missing_sheet", "missing_cell"}
        for field in fields
    )


def _summarize_fields(fields: list[WorkbookCommunicationFieldReport]) -> dict[str, int]:
    summary: dict[str, int] = {}
    for field in fields:
        summary[field.status] = summary.get(field.status, 0) + 1
    return summary


def _issues_from_fields(fields: list[WorkbookCommunicationFieldReport]) -> list[str]:
    issues: list[str] = []
    for field in fields:
        if field.issue:
            issues.append(f"{field.field}: {field.issue}")
    return issues


def _find_label_row(worksheet: Worksheet, label: str, *, max_col: int = 2) -> int | None:
    target = _normalize_label(label)
    for row in worksheet.iter_rows(max_col=max_col):
        for cell in row:
            if _normalize_label(cell.value) == target:
                return cell.row
    return None


def _find_table_value_column(worksheet: Worksheet, row_label: str, value_label: str) -> str | None:
    row_target = _normalize_label(row_label)
    value_target = _normalize_label(value_label)
    for row in worksheet.iter_rows():
        row_has_label = any(_normalize_label(cell.value) == row_target for cell in row)
        if not row_has_label:
            continue
        for cell in row:
            if _normalize_label(cell.value) == value_target:
                return cell.column_letter
    return None
