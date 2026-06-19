"""Workbook presentation fingerprint helpers.

These helpers intentionally measure broad workbook presentation structure
rather than exact cell-by-cell parity. They are used to track whether generated
models are moving toward analyst/client-ready references.
"""

from __future__ import annotations

import posixpath
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET
import zipfile

from openpyxl import load_workbook

_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS = {"main": _MAIN_NS, "rel": _REL_NS, "r": _OFFICE_REL_NS}


def workbook_presentation_fingerprint(
    path: str | Path,
    *,
    max_cells_per_sheet: int = 50_000,
) -> dict[str, Any]:
    """Return a compact structural/style fingerprint for an .xlsx workbook."""

    try:
        workbook = load_workbook(path, data_only=False, read_only=False)
    except Exception as exc:
        return _workbook_presentation_fingerprint_xml(
            path,
            max_cells_per_sheet=max_cells_per_sheet,
            parser_warning=str(exc),
        )
    visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
    sheet_fingerprints = [
        _sheet_fingerprint(sheet, max_cells_per_sheet=max_cells_per_sheet)
        for sheet in visible_sheets
    ]
    return {
        "path": str(path),
        "parser": "openpyxl",
        "sheet_count": len(workbook.worksheets),
        "visible_sheet_count": len(visible_sheets),
        "visible_sheet_names": [sheet.title for sheet in visible_sheets],
        "named_style_count": len(workbook.named_styles),
        "sheet_fingerprints": sheet_fingerprints,
        "totals": {
            "distinct_style_ids": sum(sheet["distinct_style_ids"] for sheet in sheet_fingerprints),
            "fill_color_count": sum(sheet["fill_color_count"] for sheet in sheet_fingerprints),
            "bordered_cell_count": sum(sheet["bordered_cell_count"] for sheet in sheet_fingerprints),
            "merged_range_count": sum(sheet["merged_range_count"] for sheet in sheet_fingerprints),
            "custom_width_count": sum(sheet["custom_width_count"] for sheet in sheet_fingerprints),
        },
    }


def workbook_presentation_gap(generated: dict[str, Any], reference: dict[str, Any]) -> dict[str, Any]:
    """Compare two presentation fingerprints using coarse, stable metrics."""

    generated_totals = dict(generated.get("totals") or {})
    reference_totals = dict(reference.get("totals") or {})
    total_keys = sorted(set(generated_totals) | set(reference_totals))
    return {
        "sheet_count_delta": int(generated.get("visible_sheet_count", 0))
        - int(reference.get("visible_sheet_count", 0)),
        "named_style_count_delta": int(generated.get("named_style_count", 0))
        - int(reference.get("named_style_count", 0)),
        "totals_delta": {
            key: int(generated_totals.get(key, 0)) - int(reference_totals.get(key, 0))
            for key in total_keys
        },
    }


def _sheet_fingerprint(sheet, *, max_cells_per_sheet: int) -> dict[str, Any]:
    max_row = int(sheet.max_row or 0)
    max_column = int(sheet.max_column or 0)
    style_ids: set[int] = set()
    fill_colors: set[str] = set()
    bordered_cells = 0
    scanned_cells = 0

    for row in sheet.iter_rows():
        for cell in row:
            if scanned_cells >= max_cells_per_sheet:
                break
            scanned_cells += 1
            if cell.has_style:
                style_ids.add(int(cell.style_id))
            fill_color = _fill_color(cell)
            if fill_color is not None:
                fill_colors.add(fill_color)
            if _has_border(cell):
                bordered_cells += 1
        if scanned_cells >= max_cells_per_sheet:
            break

    custom_widths = [
        dimension.width
        for dimension in sheet.column_dimensions.values()
        if dimension.width is not None
    ]
    return {
        "name": sheet.title,
        "max_row": max_row,
        "max_column": max_column,
        "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
        "distinct_style_ids": len(style_ids),
        "fill_color_count": len(fill_colors),
        "bordered_cell_count": bordered_cells,
        "merged_range_count": len(sheet.merged_cells.ranges),
        "custom_width_count": len(custom_widths),
        "scanned_cells": scanned_cells,
        "scan_truncated": scanned_cells < max_row * max_column,
    }


def _workbook_presentation_fingerprint_xml(
    path: str | Path,
    *,
    max_cells_per_sheet: int,
    parser_warning: str,
) -> dict[str, Any]:
    """Fallback fingerprint for workbooks with styles Excel tolerates but openpyxl rejects."""

    with zipfile.ZipFile(path) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relationship_targets = {
            rel.attrib.get("Id"): _resolve_workbook_target(rel.attrib.get("Target", ""))
            for rel in rels_root.findall("rel:Relationship", _NS)
        }
        style_map = _style_tables_from_archive(archive)

        sheets = []
        for sheet in workbook_root.findall("main:sheets/main:sheet", _NS):
            rel_id = sheet.attrib.get(f"{{{_OFFICE_REL_NS}}}id")
            target = relationship_targets.get(rel_id)
            sheets.append(
                {
                    "name": sheet.attrib.get("name", ""),
                    "state": sheet.attrib.get("state", "visible"),
                    "target": target,
                }
            )

        visible_sheets = [sheet for sheet in sheets if sheet["state"] == "visible"]
        sheet_fingerprints = [
            _sheet_fingerprint_xml(
                archive,
                sheet,
                style_map=style_map,
                max_cells_per_sheet=max_cells_per_sheet,
            )
            for sheet in visible_sheets
        ]

    return {
        "path": str(path),
        "parser": "xlsx_xml_fallback",
        "parser_warning": parser_warning,
        "sheet_count": len(sheets),
        "visible_sheet_count": len(visible_sheets),
        "visible_sheet_names": [sheet["name"] for sheet in visible_sheets],
        "named_style_count": style_map["named_style_count"],
        "sheet_fingerprints": sheet_fingerprints,
        "totals": {
            "distinct_style_ids": sum(sheet["distinct_style_ids"] for sheet in sheet_fingerprints),
            "fill_color_count": sum(sheet["fill_color_count"] for sheet in sheet_fingerprints),
            "bordered_cell_count": sum(sheet["bordered_cell_count"] for sheet in sheet_fingerprints),
            "merged_range_count": sum(sheet["merged_range_count"] for sheet in sheet_fingerprints),
            "custom_width_count": sum(sheet["custom_width_count"] for sheet in sheet_fingerprints),
        },
    }


def _style_tables_from_archive(archive: zipfile.ZipFile) -> dict[str, Any]:
    try:
        root = ET.fromstring(archive.read("xl/styles.xml"))
    except KeyError:
        return {"named_style_count": 0, "fills": {}, "borders": {}, "cell_xfs": []}

    fills = {
        index: _fill_color_from_xml(fill)
        for index, fill in enumerate(root.findall("main:fills/main:fill", _NS))
    }
    borders = {
        index: _bordered_from_xml(border)
        for index, border in enumerate(root.findall("main:borders/main:border", _NS))
    }
    cell_xfs = []
    for xf in root.findall("main:cellXfs/main:xf", _NS):
        cell_xfs.append(
            {
                "fill_id": _safe_int(xf.attrib.get("fillId"), 0),
                "border_id": _safe_int(xf.attrib.get("borderId"), 0),
            }
        )
    cell_styles = root.find("main:cellStyles", _NS)
    named_style_count = len(cell_styles.findall("main:cellStyle", _NS)) if cell_styles is not None else 0
    return {
        "named_style_count": named_style_count,
        "fills": fills,
        "borders": borders,
        "cell_xfs": cell_xfs,
    }


def _sheet_fingerprint_xml(
    archive: zipfile.ZipFile,
    sheet_info: dict[str, str | None],
    *,
    style_map: dict[str, Any],
    max_cells_per_sheet: int,
) -> dict[str, Any]:
    target = sheet_info.get("target")
    if not target:
        return _empty_sheet_fingerprint(str(sheet_info.get("name") or ""))

    root = ET.fromstring(archive.read(target))
    style_ids: set[int] = set()
    fill_colors: set[str] = set()
    bordered_cells = 0
    scanned_cells = 0
    max_row = 0
    max_column = 0
    truncated = False

    cells = root.findall(".//main:sheetData/main:row/main:c", _NS)
    for cell in cells:
        if scanned_cells >= max_cells_per_sheet:
            truncated = True
            break
        scanned_cells += 1
        row, column = _cell_ref_to_row_col(cell.attrib.get("r", ""))
        max_row = max(max_row, row)
        max_column = max(max_column, column)
        style_id = _safe_int(cell.attrib.get("s"), 0)
        style_ids.add(style_id)
        style = _cell_style(style_map, style_id)
        fill_color = style_map["fills"].get(style["fill_id"])
        if fill_color is not None:
            fill_colors.add(fill_color)
        if style_map["borders"].get(style["border_id"], False):
            bordered_cells += 1

    merged_ranges = root.findall("main:mergeCells/main:mergeCell", _NS)
    custom_widths = [
        col
        for col in root.findall("main:cols/main:col", _NS)
        if col.attrib.get("width") is not None and col.attrib.get("customWidth", "1") != "0"
    ]
    freeze_panes = _freeze_panes_from_xml(root)
    dimension = root.find("main:dimension", _NS)
    if dimension is not None:
        dimension_row, dimension_column = _dimension_max_row_col(dimension.attrib.get("ref", ""))
        max_row = max(max_row, dimension_row)
        max_column = max(max_column, dimension_column)

    return {
        "name": str(sheet_info.get("name") or ""),
        "max_row": max_row,
        "max_column": max_column,
        "freeze_panes": freeze_panes,
        "distinct_style_ids": len(style_ids),
        "fill_color_count": len(fill_colors),
        "bordered_cell_count": bordered_cells,
        "merged_range_count": len(merged_ranges),
        "custom_width_count": len(custom_widths),
        "scanned_cells": scanned_cells,
        "scan_truncated": truncated,
    }


def _empty_sheet_fingerprint(name: str) -> dict[str, Any]:
    return {
        "name": name,
        "max_row": 0,
        "max_column": 0,
        "freeze_panes": None,
        "distinct_style_ids": 0,
        "fill_color_count": 0,
        "bordered_cell_count": 0,
        "merged_range_count": 0,
        "custom_width_count": 0,
        "scanned_cells": 0,
        "scan_truncated": False,
    }


def _fill_color(cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return None
    color = fill.fgColor
    if color is None or color.type != "rgb" or not isinstance(color.rgb, str):
        return None
    rgb = color.rgb.strip().upper()
    if len(rgb) == 8:
        rgb = rgb[-6:]
    if len(rgb) != 6 or rgb == "000000":
        return None
    return f"#{rgb}"


def _fill_color_from_xml(fill: ET.Element) -> str | None:
    color = fill.find("main:patternFill/main:fgColor", _NS)
    if color is None:
        color = fill.find("main:fgColor", _NS)
    if color is None:
        return None
    rgb = color.attrib.get("rgb")
    if not rgb:
        return None
    return _normalize_rgb(rgb)


def _has_border(cell) -> bool:
    border = cell.border
    if border is None:
        return False
    return any(
        side is not None and side.style is not None
        for side in (border.left, border.right, border.top, border.bottom)
    )


def _bordered_from_xml(border: ET.Element) -> bool:
    return any(
        side is not None and side.attrib.get("style") is not None
        for side in [
            border.find("main:left", _NS),
            border.find("main:right", _NS),
            border.find("main:top", _NS),
            border.find("main:bottom", _NS),
        ]
    )


def _normalize_rgb(raw: str) -> str | None:
    rgb = raw.strip().upper()
    if len(rgb) == 8:
        rgb = rgb[-6:]
    if len(rgb) != 6 or rgb == "000000":
        return None
    return f"#{rgb}"


def _resolve_workbook_target(target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def _cell_style(style_map: dict[str, Any], style_id: int) -> dict[str, int]:
    cell_xfs = style_map["cell_xfs"]
    if 0 <= style_id < len(cell_xfs):
        return cell_xfs[style_id]
    return {"fill_id": 0, "border_id": 0}


def _cell_ref_to_row_col(ref: str) -> tuple[int, int]:
    letters = "".join(ch for ch in ref if ch.isalpha())
    digits = "".join(ch for ch in ref if ch.isdigit())
    column = 0
    for char in letters.upper():
        column = column * 26 + (ord(char) - ord("A") + 1)
    return (_safe_int(digits, 0), column)


def _dimension_max_row_col(ref: str) -> tuple[int, int]:
    if ":" in ref:
        ref = ref.split(":", 1)[1]
    return _cell_ref_to_row_col(ref)


def _freeze_panes_from_xml(root: ET.Element) -> str | None:
    pane = root.find("main:sheetViews/main:sheetView/main:pane", _NS)
    if pane is None:
        return None
    return pane.attrib.get("topLeftCell")


def _safe_int(value: str | None, default: int) -> int:
    try:
        return int(value) if value is not None else default
    except ValueError:
        return default


__all__ = [
    "workbook_presentation_fingerprint",
    "workbook_presentation_gap",
]
