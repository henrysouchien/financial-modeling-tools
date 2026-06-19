from __future__ import annotations

from datetime import datetime, timezone
import json
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import load_workbook

from .driver_resolver import resolve_driver_cells, resolve_driver_key
from .model_build_context import ModelBuildContext
from .models import FinancialModel
from .reader import read_model
from .tools import clear_cache

if TYPE_CHECKING:
  from .business_model_compiler import CompiledDriverRegistry

try:
  from research.repository import get_repository_factory
except ModuleNotFoundError:  # pragma: no cover - fallback for package imports
  from api.research.repository import get_repository_factory


RESEARCH_CONTEXT_SHEET = "_ResearchContext"


def _utc_now_iso() -> str:
  return datetime.now(timezone.utc).isoformat()


def _artifact_from_handoff(handoff_row: dict[str, Any]) -> dict[str, Any]:
  artifact = handoff_row.get("artifact")
  if isinstance(artifact, dict):
    return artifact
  if isinstance(artifact, str) and artifact.strip():
    parsed = json.loads(artifact)
    if isinstance(parsed, dict):
      return parsed
  raise ValueError(f"handoff {handoff_row.get('id')} has invalid artifact payload")


def _clear_sheet(worksheet) -> None:
  if worksheet.max_row > 0:
    worksheet.delete_rows(1, worksheet.max_row)


def _load_annotation_context(
  model_path: str,
  handoff_id: int,
  user_id: int,
  model: FinancialModel | None = None,
) -> tuple[Path, dict[str, Any], dict[str, Any], Any, FinancialModel]:
  resolved_model_path = Path(model_path).expanduser()
  if not resolved_model_path.is_file():
    raise FileNotFoundError(f"Model file not found: {resolved_model_path}")

  repo = get_repository_factory().get(user_id)
  handoff = repo.get_handoff(handoff_id)
  if handoff is None:
    raise ValueError(f"handoff not found: {handoff_id}")
  artifact = _artifact_from_handoff(handoff)

  workbook = load_workbook(resolved_model_path, data_only=False)
  model_obj = model
  if model_obj is None:
    loaded = read_model(str(resolved_model_path), mode="full")
    if not isinstance(loaded, FinancialModel):
      raise TypeError("read_model did not return a FinancialModel")
    model_obj = loaded
  if not model_obj._index:
    model_obj.build_index()

  return resolved_model_path, handoff, artifact, workbook, model_obj


def _write_context_sheet(workbook, handoff: dict[str, Any], artifact: dict[str, Any]) -> None:
  context_sheet = workbook[RESEARCH_CONTEXT_SHEET] if RESEARCH_CONTEXT_SHEET in workbook.sheetnames else workbook.create_sheet(
    RESEARCH_CONTEXT_SHEET
  )
  _clear_sheet(context_sheet)
  context_sheet.sheet_state = "hidden"
  context_sheet["A1"] = "field"
  context_sheet["B1"] = "value"
  context_sheet["A2"] = "handoff_id"
  context_sheet["B2"] = int(handoff["id"])
  context_sheet["A3"] = "research_file_id"
  context_sheet["B3"] = int(handoff["research_file_id"])
  context_sheet["A4"] = "ticker"
  context_sheet["B4"] = handoff["ticker"]
  context_sheet["A5"] = "version"
  context_sheet["B5"] = int(handoff["version"])
  context_sheet["A6"] = "status"
  context_sheet["B6"] = handoff["status"]
  context_sheet["A7"] = "annotated_at"
  context_sheet["B7"] = _utc_now_iso()
  context_sheet["A8"] = "artifact_json"
  context_sheet["B8"] = json.dumps(artifact, sort_keys=True)


def _finalize_annotation(workbook, resolved_model_path: Path) -> dict[str, Any]:
  workbook.calculation.fullCalcOnLoad = True
  workbook.calculation.forceFullCalc = True
  workbook.save(resolved_model_path)
  clear_cache()
  return {
    "model_path": str(resolved_model_path),
    "annotated_at": _utc_now_iso(),
  }


def _should_skip_annotation(
  item_id: str,
  compiled_registry: "CompiledDriverRegistry | None",
  model: FinancialModel,
) -> bool:
  if compiled_registry is None:
    return False

  bm_item_ids = set(compiled_registry.node_items.values()) | set(compiled_registry.driver_keys.values())
  if item_id not in bm_item_ids:
    return False

  try:
    item = model.get_item(item_id)
  except KeyError:
    return False
  return item.projected is not None


def annotate_model_with_research(
  model_path: str,
  handoff_id: int,
  user_id: int,
  model: FinancialModel | None = None,
  compiled_registry: "CompiledDriverRegistry | None" = None,
) -> dict[str, Any]:
  resolved_model_path, handoff, artifact, workbook, model_obj = _load_annotation_context(
    model_path,
    handoff_id,
    user_id,
    model=model,
  )

  assumptions_written = 0
  assumptions_skipped: list[dict[str, str]] = []
  for assumption in artifact.get("assumptions") or []:
    if not isinstance(assumption, dict):
      continue
    driver_key = str(assumption.get("driver") or "").strip()
    if not driver_key:
      assumptions_skipped.append({"driver": "", "reason": "missing driver"})
      continue
    if assumption.get("value") is None:
      assumptions_skipped.append({"driver": driver_key, "reason": "missing value"})
      continue
    try:
      if compiled_registry is None:
        cells = resolve_driver_cells(model_obj, driver_key)
      else:
        item_id = resolve_driver_key(driver_key, compiled_registry=compiled_registry)
        if _should_skip_annotation(item_id, compiled_registry, model_obj):
          continue
        cells = resolve_driver_cells(model_obj, driver_key, compiled_registry=compiled_registry)
    except Exception as exc:
      assumptions_skipped.append({"driver": driver_key, "reason": str(exc)})
      continue

    for sheet_name, cell_address, _period in cells:
      workbook[sheet_name][cell_address] = assumption.get("value")
      assumptions_written += 1

  _write_context_sheet(workbook, handoff, artifact)
  result = _finalize_annotation(workbook, resolved_model_path)

  return {
    **result,
    "assumptions_written": assumptions_written,
    "assumptions_skipped": assumptions_skipped,
  }


def annotate_model_with_research_from_mbc(
  mbc: ModelBuildContext,
  model_path: str,
  handoff_id: int,
  user_id: int,
  *,
  model: FinancialModel | None = None,
  compiled_registry: "CompiledDriverRegistry | None" = None,
) -> dict[str, Any]:
  resolved_model_path, handoff, artifact, workbook, model_obj = _load_annotation_context(
    model_path,
    handoff_id,
    user_id,
    model=model,
  )

  assumptions_written = 0
  for driver_key, driver in mbc.drivers.items():
    if compiled_registry is None:
      cells = resolve_driver_cells(model_obj, driver_key, periods=driver.periods)
    else:
      item_id = resolve_driver_key(driver_key, compiled_registry=compiled_registry)
      if _should_skip_annotation(item_id, compiled_registry, model_obj):
        continue
      cells = resolve_driver_cells(
        model_obj,
        driver_key,
        periods=driver.periods,
        compiled_registry=compiled_registry,
      )
    for sheet_name, cell_address, _period in cells:
      workbook[sheet_name][cell_address] = driver.value
      assumptions_written += 1

  _write_context_sheet(workbook, handoff, artifact)
  result = _finalize_annotation(workbook, resolved_model_path)
  return {
    **result,
    "assumptions_written": assumptions_written,
    "assumptions_skipped": [],
  }


__all__ = ["RESEARCH_CONTEXT_SHEET", "annotate_model_with_research", "annotate_model_with_research_from_mbc"]
