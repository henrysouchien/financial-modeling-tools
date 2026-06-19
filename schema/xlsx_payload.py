"""Serialize an existing .xlsx workbook into an apply_render_plan payload."""

from __future__ import annotations

from collections import defaultdict
from itertools import groupby
import os
import string
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import (
  column_index_from_string,
  coordinate_from_string,
  get_column_letter,
  range_boundaries,
)

from .renderer import CellFormat, CellWrite, _batch_sheet_formats, _batch_sheet_writes


_VALID_CONFLICT_STRATEGIES = {"fail_on_collision", "overwrite"}
_HEX_CHARS = set(string.hexdigits)
_BATCHING_DISABLED_FALSEY = {"", "0", "false", "no", "off"}


def xlsx_to_addin_payload(
  workbook_path: str,
  conflict_strategy: str = "fail_on_collision",
) -> dict[str, Any]:
  """Walk a workbook from disk into the apply_render_plan tool input shape.

  Set XLSX_PAYLOAD_DISABLE_BATCHING to a truthy value to emit the v1.5-style
  per-cell payload for live debugging.
  """
  if conflict_strategy not in _VALID_CONFLICT_STRATEGIES:
    raise ValueError(
      "conflict_strategy must be 'fail_on_collision' or 'overwrite'; "
      f"got {conflict_strategy!r}"
    )

  sheets: list[dict[str, Any]] = []
  walked_writes: list[CellWrite] = []
  walked_formats: list[CellFormat] = []

  workbook = load_workbook(Path(workbook_path), data_only=False)
  try:
    for worksheet in workbook.worksheets:
      sheet_payload: dict[str, Any] = {"name": worksheet.title}
      column_widths = {
        column: float(dimension.width)
        for column, dimension in worksheet.column_dimensions.items()
        if dimension.width is not None
      }
      if column_widths:
        sheet_payload["column_widths"] = column_widths
      if worksheet.freeze_panes:
        sheet_payload["freeze_panes"] = str(worksheet.freeze_panes)
      sheets.append(sheet_payload)
      for cell in worksheet.iter_rows():
        for current in cell:
          if current.value is not None:
            walked_writes.append(
              CellWrite(
                sheet=worksheet.title,
                cell=current.coordinate,
                value=current.value,
              )
            )

          cell_format = _extract_format(current)
          if cell_format:
            walked_formats.append(
              CellFormat(
                sheet=worksheet.title,
                range=f"{current.coordinate}:{current.coordinate}",
                **cell_format,
              )
            )
  finally:
    workbook.close()

  sheet_names = [sheet["name"] for sheet in sheets]
  if _batching_disabled():
    writes = _per_cell_writes(walked_writes)
    formats = _per_cell_formats(walked_formats)
  else:
    writes = _batch_writes_by_sheet(walked_writes, sheet_names)
    formats = _batch_formats_by_sheet(walked_formats, sheet_names)

  return {
    "sheets": sheets,
    "writes": writes,
    "formats": formats,
    "conflict_strategy": conflict_strategy,
  }


def _batching_disabled() -> bool:
  value = os.getenv("XLSX_PAYLOAD_DISABLE_BATCHING", "").strip().lower()
  return value not in _BATCHING_DISABLED_FALSEY


def _per_cell_writes(writes: list[CellWrite]) -> list[dict[str, Any]]:
  return [
    {
      "sheet": write.sheet,
      "range": write.cell,
      "values": write.value,
    }
    for write in writes
  ]


def _per_cell_formats(formats: list[CellFormat]) -> list[dict[str, Any]]:
  return [{"sheet": cell_format.sheet, **_format_payload(cell_format)} for cell_format in formats]


def _batch_writes_by_sheet(
  writes: list[CellWrite],
  sheet_names: list[str],
) -> list[dict[str, Any]]:
  grouped: defaultdict[str, list[CellWrite]] = defaultdict(list)
  for write in writes:
    grouped[write.sheet].append(write)

  payloads: list[dict[str, Any]] = []
  for sheet_name in sheet_names:
    sheet_writes = sorted(grouped.get(sheet_name, []), key=_write_sort_key)
    for _row, row_group in groupby(sheet_writes, key=_write_row):
      row_writes = list(row_group)
      cols = [_write_col_index(write) for write in row_writes]
      for payload in _batch_sheet_writes(row_writes, min(cols), max(cols)):
        payloads.append({"sheet": sheet_name, **payload})
  return payloads


def _coalesce_row_formats(formats: list[CellFormat]) -> list[CellFormat]:
  grouped: defaultdict[tuple[str, int, tuple[Any, ...]], list[tuple[int, CellFormat]]] = defaultdict(list)
  for cell_format in formats:
    min_col, min_row, max_col, max_row = range_boundaries(cell_format.range)
    for row in range(min_row, max_row + 1):
      for col in range(min_col, max_col + 1):
        grouped[(cell_format.sheet, row, _format_key(cell_format))].append((col, cell_format))

  coalesced: list[CellFormat] = []
  for (sheet_name, row, _format_key_value), cells in grouped.items():
    ordered = sorted(cells, key=lambda item: item[0])
    if not ordered:
      continue

    run_start_col, run_format = ordered[0]
    run_end_col = run_start_col
    for col, cell_format in ordered[1:]:
      if col == run_end_col + 1:
        run_end_col = col
        continue

      coalesced.append(_row_format(sheet_name, row, run_start_col, run_end_col, run_format))
      run_start_col = col
      run_end_col = col
      run_format = cell_format

    coalesced.append(_row_format(sheet_name, row, run_start_col, run_end_col, run_format))

  return coalesced


def _batch_formats_by_sheet(
  formats: list[CellFormat],
  sheet_names: list[str],
) -> list[dict[str, Any]]:
  coalesced = _coalesce_row_formats(formats)
  grouped: defaultdict[str, list[CellFormat]] = defaultdict(list)
  for cell_format in coalesced:
    grouped[cell_format.sheet].append(cell_format)

  payloads: list[dict[str, Any]] = []
  for sheet_name in sheet_names:
    sheet_formats = sorted(grouped.get(sheet_name, []), key=_format_range_sort_key)
    for payload in _batch_sheet_formats(sheet_formats):
      payloads.append({"sheet": sheet_name, **payload})
  return payloads


def _row_format(
  sheet_name: str,
  row: int,
  start_col: int,
  end_col: int,
  cell_format: CellFormat,
) -> CellFormat:
  return CellFormat(
    sheet=sheet_name,
    range=f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}",
    bold=cell_format.bold,
    font_color=cell_format.font_color,
    number_format=cell_format.number_format,
    underline=cell_format.underline,
    fill_color=cell_format.fill_color,
    top_border_color=cell_format.top_border_color,
    bottom_border_color=cell_format.bottom_border_color,
  )


def _write_sort_key(write: CellWrite) -> tuple[int, int]:
  col, row = coordinate_from_string(write.cell)
  return row, column_index_from_string(col)


def _write_row(write: CellWrite) -> int:
  return _write_sort_key(write)[0]


def _write_col_index(write: CellWrite) -> int:
  return _write_sort_key(write)[1]


def _format_range_sort_key(cell_format: CellFormat) -> tuple[int, int, int, int]:
  min_col, min_row, max_col, max_row = range_boundaries(cell_format.range)
  return min_col, max_col, min_row, max_row


def _format_key(cell_format: CellFormat) -> tuple[Any, ...]:
  return (
    cell_format.bold,
    cell_format.font_color.lower() if cell_format.font_color else None,
    cell_format.number_format,
    cell_format.underline,
    cell_format.fill_color.lower() if cell_format.fill_color else None,
    cell_format.top_border_color.lower() if cell_format.top_border_color else None,
    cell_format.bottom_border_color.lower() if cell_format.bottom_border_color else None,
  )


def _format_payload(cell_format: CellFormat) -> dict[str, Any]:
  payload: dict[str, Any] = {"range": cell_format.range}
  if cell_format.bold is not None:
    payload["bold"] = cell_format.bold
  if cell_format.font_color is not None:
    payload["font_color"] = cell_format.font_color
  if cell_format.number_format is not None:
    payload["number_format"] = cell_format.number_format
  if cell_format.underline is not None:
    payload["underline"] = cell_format.underline
  if cell_format.fill_color is not None:
    payload["fill_color"] = cell_format.fill_color
  if cell_format.top_border_color is not None:
    payload["top_border_color"] = cell_format.top_border_color
  if cell_format.bottom_border_color is not None:
    payload["bottom_border_color"] = cell_format.bottom_border_color
  return payload


def _extract_format(cell) -> dict[str, Any]:
  cell_format: dict[str, Any] = {}
  font = cell.font

  if font.bold is True:
    cell_format["bold"] = True

  font_color = _extract_rgb_font_color(font.color)
  if font_color is not None:
    cell_format["font_color"] = font_color

  if cell.number_format != "General":
    cell_format["number_format"] = cell.number_format

  if font.underline is not None and str(font.underline).lower() != "none":
    cell_format["underline"] = True

  fill_color = _extract_rgb_color(cell.fill.fgColor)
  if cell.fill.fill_type is not None and fill_color is not None and fill_color != "#000000":
    cell_format["fill_color"] = fill_color

  top_border_color = _extract_border_color(cell.border.top)
  if top_border_color is not None:
    cell_format["top_border_color"] = top_border_color

  bottom_border_color = _extract_border_color(cell.border.bottom)
  if bottom_border_color is not None:
    cell_format["bottom_border_color"] = bottom_border_color

  return cell_format


def _extract_rgb_font_color(color) -> str | None:
  return _extract_rgb_color(color)


def _extract_rgb_color(color) -> str | None:
  if color is None or getattr(color, "type", None) != "rgb":
    return None

  rgb = getattr(color, "rgb", None)
  if not isinstance(rgb, str):
    return None

  value = rgb.strip()
  if len(value) not in {6, 8}:
    return None
  if any(ch not in _HEX_CHARS for ch in value):
    return None
  return f"#{value[-6:]}"


def _extract_border_color(side) -> str | None:
  if side is None or side.style is None:
    return None
  return _extract_rgb_color(side.color) or "#000000"
